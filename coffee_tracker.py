from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os


def get_previous_month_file():
    current_date = datetime.now()
    # Go back to the first day of the current month -- that is if you are running in the first day of the current month
    first_of_current_month = current_date.replace(day=1)
    # Then go back one more day to get to the previous month
    last_of_previous_month = first_of_current_month - timedelta(days=1)
    # Format the filename using this date
    previous_month_str = last_of_previous_month.strftime("%B_%Y")

    filename = f"Coffee_Sheet_{previous_month_str}.xlsx"
    return filename if os.path.exists(filename) else None


def load_previous_data(filename):
    wb = load_workbook(filename)
    ws = wb.active
    employees = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, to_pay, paid_last_month, _ = row
        employees[name] = {
            'coffees': 0,
            'unpaid_amount': to_pay if paid_last_month == "No" else 0,
            'paid_last_month': paid_last_month == "Yes"
        }
    return employees


def update_employee_data(employees):
    for name in list(employees.keys()):
        print(f"\nUpdating data for {name}")
        coffees = int(input(f"How many coffees did {name} have this month? "))
        paid = input(f"Did {name} pay last month? (y/n): ").lower() == 'y'
        employees[name]['coffees'] = coffees
        employees[name]['paid_last_month'] = paid
        if paid:
            employees[name]['unpaid_amount'] = 0

    while True:
        add_employee = input("\nDo you want to add a new employee? (y/n): ").lower()
        if add_employee != 'y':
            break
        name = input("Enter the name of the new employee: ")
        coffees = int(input(f"How many coffees did {name} have? "))
        paid = input(f"Did {name} pay last month? (y/n): ").lower() == 'y'
        employees[name] = {'coffees': coffees, 'unpaid_amount': 0, 'paid_last_month': paid}

    while True:
        remove_employee = input("\nDo you want to remove an employee? (y/n): ").lower()
        if remove_employee != 'y':
            break
        name = input("Enter the name of the employee to remove: ")
        if name in employees:
            del employees[name]
            print(f"{name} has been removed.")
        else:
            print(f"{name} not found in the list.")

    return employees


def calculate_price(coffees):
    return coffees * 7.50


def generate_sheet(employees):
    wb = Workbook()
    ws = wb.active
    current_date = datetime.now()
    first_of_current_month = current_date.replace(day=1)
    last_of_previous_month = first_of_current_month - timedelta(days=1)

    ws.title = last_of_previous_month.strftime("%B %Y")

    ws['A1'] = "Name"
    ws['B1'] = "To pay"
    ws['C1'] = "Paid?"
    ws['D1'] = "Coffees"

    for idx, (name, data) in enumerate(employees.items(), start=2):
        current_month_price = calculate_price(data['coffees'])
        total_to_pay = current_month_price + data['unpaid_amount']
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=total_to_pay)
        ws.cell(row=idx, column=3, value="Yes" if data['paid_last_month'] else "No")
        ws.cell(row=idx, column=4, value=data['coffees'])

    filename = f"Coffee_Sheet_{last_of_previous_month.strftime('%B_%Y')}.xlsx"
    wb.save(filename)
    print(f"Sheet saved as {filename}")


def main():
    previous_file = get_previous_month_file()
    if previous_file:
        employees = load_previous_data(previous_file)
        print(f"Loaded data from {previous_file}")
    else:
        employees = {}
        print("No previous month's data found. Starting fresh.")

    employees = update_employee_data(employees)
    generate_sheet(employees)


if __name__ == "__main__":
    main()